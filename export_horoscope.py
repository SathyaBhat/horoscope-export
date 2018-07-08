#!/usr/bin/env python

from __future__ import unicode_literals
from openpyxl import load_workbook
from sys import exit
from pprint import pprint

from logging import basicConfig, getLogger, DEBUG

import pymysql
import argparse

def create_db_connection(args):
    try:
        connection = pymysql.connect(host=args.host,
                                    user=args.user,
                                    passwd=args.password,
                                    db=args.database,
                                    use_unicode=True, 
                                    charset="utf8")
    except pymysql.err.InternalError as e: 
        code, msg = e.args
        if code == 1049 and msg == "Unknown database '{}'".format(args.database):
            connection = pymysql.connect(host=args.host,
                                    user=args.user,
                                    passwd=args.password,
                                    use_unicode=True, 
                                    charset="utf8")
            connection.cursor().execute('create database {}'.format(args.database))
            connection = pymysql.connect(host=args.host,
                                    user=args.user,
                                    passwd=args.password,
                                    db=args.database,
                                    use_unicode=True, 
                                    charset="utf8") 
    return connection

def create_database(connection):
    try:
        connection.cursor().execute('create database horoscope')
    except pymysql.err.ProgrammingError as e:
        code, msg = e.args
        if code == 1007:
            log.info("Database exists, not recreating.")
        else:
            log.error("Code: {} \n Messsage: {}".format(code, msg))
            exit(-1)

def create_tables(connection, table_name, headers):
    log.debug("table_name name: {} \n headers: {}".format(table_name, headers)) 
    sql_query = 'CREATE TABLE `{0}` ( \
                                `{1}` bigint(11) unsigned NOT NULL AUTO_INCREMENT, \
                                `{2}` varchar(11) CHARACTER SET utf8 DEFAULT NULL, \
                                `{3}` mediumtext CHARACTER SET utf8,  \
                                `{4}` text CHARACTER SET utf8,  \
                                `{5}` int(11) DEFAULT NULL,  \
                                `{6}` varchar(48) CHARACTER SET utf8 DEFAULT NULL, \
                                `{7}` mediumtext CHARACTER SET utf8,  \
                                PRIMARY KEY (`{1}`) \
                                ) ENGINE=InnoDB'.format(table_name,
                                                        *headers)
    log.debug('SQL Query:\n {}'.format(sql_query))
    try:
        connection.cursor().execute(sql_query)
    except pymysql.err.InternalError as e:
        code, msg = e.args
        if code == 1050:
            log.info("Table exists, not creating")
        else:
            log.error("Code: {} \n Messsage: {}".format(code, msg))
            exit(-1)

def insert_data_to_tables(connection, table, headers, data):
    log.info("Inserting data to tables")
    log.debug("Data: \n {}".format(data))
    headers = headers[1:]
    sql_query = "INSERT INTO {}({}, {}, {}, {}, {}, {}) \
                    VALUES     (%s, %s, %s, %s, %s, %s)".format(table, *headers)
    args = (data[1].value.strftime('%Y-%m-%d'),data[2].value,data[3].value,data[4].value,data[5].value,data[6].value )
    log.debug("SQL Query {}".format(sql_query))
    connection.cursor().execute(sql_query, args)

def export_to_mysql(file_name, connection):
    book = load_workbook(file_name)
    sheets = book.sheetnames
    values = []

    for sheet in sheets:
        current_sheet = book.get_sheet_by_name(sheet)
        cells = current_sheet['A1': 'G32']
        headers = []
        
        for header in cells[0]:
            headers.append(header.value)
        create_tables(connection, sheet, headers)
        for data in cells[1:]:
            try:
                insert_data_to_tables(connection, sheet, headers, data)
                connection.commit()
            except (pymysql.err.InternalError, pymysql.err.DataError) as e:
                code, msg = e.args
                log.error("Code: {} \n Messsage: {}".format(code, msg))
                log.debug("Sheet: {} \n Data {}".format(sheet, data))
                connection.rollback()
                exit(-1)

if __name__=='__main__':
    basicConfig(level=DEBUG,
                    format='%(levelname)s: %(asctime)s -'
                    ' %(funcName)s - %(message)s')
    log = getLogger('horo')
    parser = argparse.ArgumentParser(description='Script to parse the horoscope Excel.')
    parser.add_argument('--file', '-f', required=True, help='Filename to be read')
    parser.add_argument('--host', '-o', required=True, help='MySQL Host')
    parser.add_argument('--user', '-u', required=True, help='MySQL Username')
    parser.add_argument('--password', '-p', required=True, help='MySQL password')
    parser.add_argument('--database', '-d', required=True, help='MySQL database to connect to')
    args = parser.parse_args()
    connection = create_db_connection(args)
    #create_database(connection)
    export_to_mysql(args.file, connection)
